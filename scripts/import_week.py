#!/usr/bin/env python3
"""
One-time importer: push a weekly activities spreadsheet into Habit Quest.

It signs in with a Habit Quest account (parent or child) using the Firebase
Auth REST API, resolves the child's family/member, parses the timetable
spreadsheet, and writes (or updates) the plan document for a given week
directly in Firestore via the REST API.

No service account needed — it uses the same email/password login the app
uses, so Firestore security rules apply normally.

Usage (PowerShell):
  $env:HQ_EMAIL    = "parent-or-child@example.com"
  $env:HQ_PASSWORD = "the-password"
  # optional, only if signing in as a parent with more than one child:
  $env:HQ_CHILD_EMAIL = "child@example.com"
  # optional, defaults to the current week's Monday:
  $env:HQ_WEEK_START  = "2026-04-20"
  python scripts/import_week.py "C:\\path\\to\\Kids_Daily_Activities_Timesheet.xlsx"

If HQ_EMAIL / HQ_PASSWORD are not set, the script prompts for them.
"""

import os
import re
import sys
import json
import getpass
from datetime import datetime, date, timedelta, timezone

import requests
import openpyxl


def _config_from_options():
    """Pull the Firebase web apiKey/projectId from lib/firebase_options.dart.

    This is a public Firebase *client* config (safe to live in the app source),
    so we read it from there instead of duplicating the key inside this script.
    """
    path = os.path.join(os.path.dirname(__file__), "..", "lib",
                        "firebase_options.dart")
    api_key = project_id = None
    try:
        with open(path, "r", encoding="utf-8") as f:
            text = f.read()
        m = re.search(r"apiKey:\s*'([^']+)'", text)
        if m:
            api_key = m.group(1)
        m = re.search(r"projectId:\s*'([^']+)'", text)
        if m:
            project_id = m.group(1)
    except OSError:
        pass
    return api_key, project_id


_OPT_KEY, _OPT_PID = _config_from_options()

# Prefer env vars, fall back to the app's firebase_options.dart.
API_KEY = os.environ.get("HQ_API_KEY") or _OPT_KEY
PROJECT_ID = os.environ.get("HQ_PROJECT_ID") or _OPT_PID

if not API_KEY or not PROJECT_ID:
    sys.exit(
        "Could not determine the Firebase web API key / project id.\n"
        "Set HQ_API_KEY and HQ_PROJECT_ID, or run from the repo root so "
        "lib/firebase_options.dart can be read."
    )

AUTH_BASE = "https://identitytoolkit.googleapis.com/v1"
FS_BASE = f"https://firestore.googleapis.com/v1/projects/{PROJECT_ID}/databases/(default)/documents"

DAY_NAMES = [
    "monday", "tuesday", "wednesday", "thursday",
    "friday", "saturday", "sunday",
]
DAY_SHEETS = {
    "monday": "Monday", "tuesday": "Tuesday", "wednesday": "Wednesday",
    "thursday": "Thursday", "friday": "Friday", "saturday": "Saturday",
    "sunday": "Sunday",
}

# Activity title -> app category name (must match TaskCategory enum).
CATEGORY_MAP = {
    # study
    "Do Math": "study",
    "Read English": "study",
    "Read Tamil": "study",
    "Reading Books": "study",
    # creative
    "Draw": "creative",
    "Making Crafts": "creative",
    "Doing Lego": "creative",
    "Robot Kit": "creative",
    "Human Being Build": "creative",
    # exercise
    "Pickleball": "exercise",
    "Table Tennis (TT)": "exercise",
    "Basketball": "exercise",
    "Ball Catching": "exercise",
    # screen time (unhealthy / digital)
    "Watch TV": "screenTime",
    "iPad Games": "screenTime",
    # social / family
    "Card / Board Games with Family": "social",
    "Dinner with Family": "social",
    # everything else (meals / rest) -> custom, keeps its own label
}
HEALTHY_BY_DEFAULT = {
    "exercise": True, "study": True, "chores": True, "creative": True,
    "social": True, "screenTime": False, "sleep": True, "custom": True,
}


def category_for(title: str) -> str:
    return CATEGORY_MAP.get(title.strip(), "custom")


# ---------- Firestore REST value helpers ----------

def to_value(v):
    if v is None:
        return {"nullValue": None}
    if isinstance(v, bool):
        return {"booleanValue": v}
    if isinstance(v, int):
        return {"integerValue": str(v)}
    if isinstance(v, str):
        return {"stringValue": v}
    if isinstance(v, list):
        return {"arrayValue": {"values": [to_value(x) for x in v]}}
    if isinstance(v, dict):
        return {"mapValue": {"fields": {k: to_value(x) for k, x in v.items()}}}
    raise TypeError(f"Unsupported value type: {type(v)}")


def to_fields(d: dict) -> dict:
    return {k: to_value(v) for k, v in d.items()}


def from_value(val: dict):
    if "nullValue" in val:
        return None
    if "booleanValue" in val:
        return val["booleanValue"]
    if "integerValue" in val:
        return int(val["integerValue"])
    if "stringValue" in val:
        return val["stringValue"]
    if "arrayValue" in val:
        return [from_value(x) for x in val["arrayValue"].get("values", [])]
    if "mapValue" in val:
        return {k: from_value(x)
                for k, x in val["mapValue"].get("fields", {}).items()}
    if "timestampValue" in val:
        return val["timestampValue"]
    return None


# ---------- Auth ----------

def sign_in(email: str, password: str):
    r = requests.post(
        f"{AUTH_BASE}/accounts:signInWithPassword?key={API_KEY}",
        json={"email": email, "password": password, "returnSecureToken": True},
        timeout=30,
    )
    if r.status_code != 200:
        msg = r.json().get("error", {}).get("message", r.text)
        sys.exit(f"Login failed: {msg}")
    data = r.json()
    return data["idToken"], data["localId"]


def auth_headers(token: str):
    return {"Authorization": f"Bearer {token}"}


# ---------- Firestore reads ----------

def get_doc(path: str, token: str):
    r = requests.get(f"{FS_BASE}/{path}", headers=auth_headers(token), timeout=30)
    if r.status_code == 404:
        return None
    r.raise_for_status()
    return {k: from_value(v) for k, v in r.json().get("fields", {}).items()}


def run_query(parent_path: str, collection_id: str, where, token: str):
    """where = list of (field, value) ANDed with EQUAL."""
    filters = [{
        "fieldFilter": {
            "field": {"fieldPath": f},
            "op": "EQUAL",
            "value": to_value(v),
        }
    } for f, v in where]
    body = {
        "structuredQuery": {
            "from": [{"collectionId": collection_id}],
            "where": {"compositeFilter": {"op": "AND", "filters": filters}},
            "limit": 5,
        }
    }
    url = f"{FS_BASE}/{parent_path}:runQuery" if parent_path else f"{FS_BASE}:runQuery"
    r = requests.post(url, headers=auth_headers(token), json=body, timeout=30)
    r.raise_for_status()
    out = []
    for row in r.json():
        doc = row.get("document")
        if not doc:
            continue
        fields = {k: from_value(v) for k, v in doc.get("fields", {}).items()}
        out.append((doc["name"], fields))
    return out


def list_collection(parent_path: str, collection_id: str, token: str):
    url = f"{FS_BASE}/{parent_path}/{collection_id}" if parent_path else f"{FS_BASE}/{collection_id}"
    r = requests.get(url, headers=auth_headers(token), timeout=30)
    r.raise_for_status()
    out = []
    for doc in r.json().get("documents", []):
        fields = {k: from_value(v) for k, v in doc.get("fields", {}).items()}
        out.append((doc["name"], fields))
    return out


# ---------- Spreadsheet parsing ----------

def parse_time(t: str) -> int:
    """'8:30 AM' -> minutes since midnight."""
    dt = datetime.strptime(t.strip(), "%I:%M %p")
    return dt.hour * 60 + dt.minute


def parse_day(ws):
    """Return list of merged tasks for a single day sheet."""
    raw = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        time_range, activity = row[0], row[1]
        if not time_range or not activity:
            continue
        if "-" not in str(time_range):
            continue
        start_s, end_s = [x.strip() for x in str(time_range).split("-", 1)]
        start = parse_time(start_s)
        end = parse_time(end_s)
        raw.append((start, end - start, str(activity).strip()))

    # Merge consecutive identical activities into one longer task.
    merged = []
    for start, dur, title in raw:
        if merged and merged[-1]["title"] == title and \
                merged[-1]["hour"] * 60 + merged[-1]["minute"] + merged[-1]["duration"] == start:
            merged[-1]["duration"] += dur
        else:
            merged.append({
                "hour": start // 60,
                "minute": start % 60,
                "duration": dur,
                "title": title,
            })

    tasks = []
    for i, m in enumerate(merged):
        cat = category_for(m["title"])
        tasks.append({
            "taskId": f"imp-{ws.title.lower()}-{i:02d}-{int(datetime.now().timestamp())}",
            "hour": m["hour"],
            "minute": m["minute"],
            "duration": m["duration"],
            "title": m["title"],
            "category": cat,
            "customCategoryName": m["title"] if cat == "custom" else None,
            "isDigitalActivity": cat == "screenTime",
            "isHealthy": HEALTHY_BY_DEFAULT[cat],
        })
    return tasks


def build_days(xlsx_path: str):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    days = {}
    for key, sheet in DAY_SHEETS.items():
        if sheet in wb.sheetnames:
            days[key] = parse_day(wb[sheet])
        else:
            days[key] = []
    return days


# ---------- Main ----------

def main():
    if len(sys.argv) < 2:
        sys.exit("Usage: python scripts/import_week.py <path-to-xlsx>")
    xlsx_path = sys.argv[1]
    if not os.path.exists(xlsx_path):
        sys.exit(f"File not found: {xlsx_path}")

    email = os.environ.get("HQ_EMAIL") or input("Habit Quest email: ").strip()
    password = os.environ.get("HQ_PASSWORD") or getpass.getpass("Password: ")

    week_start = os.environ.get("HQ_WEEK_START")
    if not week_start:
        today = date.today()
        monday = today - timedelta(days=today.weekday())
        week_start = monday.isoformat()

    print(f"Signing in as {email} ...")
    token, uid = sign_in(email, password)

    profile = get_doc(f"userProfiles/{uid}", token)
    if not profile:
        sys.exit("No userProfile found for this account. Log into the app once, "
                 "then retry.")
    family_id = profile.get("familyId")
    role = profile.get("role")
    member_id = profile.get("memberId")
    if not family_id:
        sys.exit("Account has no family.")

    # If signed in as a parent, resolve which child the plan is for.
    if role == "parent":
        child_email = os.environ.get("HQ_CHILD_EMAIL")
        members = list_collection(f"families/{family_id}", "members", token)
        children = [(name.split("/")[-1], f) for name, f in members
                    if f.get("role") == "child"]
        if not children:
            sys.exit("No child members found in this family. Add a child first.")
        if child_email:
            match = [(mid, f) for mid, f in children
                     if (f.get("email") or "").lower() == child_email.lower()]
            if not match:
                sys.exit(f"No child with email {child_email}.")
            member_id = match[0][0]
        elif len(children) == 1:
            member_id = children[0][0]
            print(f"Importing for child: {children[0][1].get('displayName')}")
        else:
            names = ", ".join(
                f"{f.get('displayName')} <{f.get('email')}>" for _, f in children)
            sys.exit("Multiple children found. Set HQ_CHILD_EMAIL to one of: "
                     + names)

    if not member_id:
        sys.exit("Could not resolve memberId.")

    print(f"Family: {family_id}  Member: {member_id}  Week: {week_start}")

    days = build_days(xlsx_path)
    total = sum(len(v) for v in days.values())
    print(f"Parsed {total} tasks across 7 days.")

    now_iso = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
    plan_fields = to_fields({
        "memberId": member_id,
        "weekStart": week_start,
        "status": "approved",      # parent is importing, so pre-approved
        "parentNote": None,
        "days": days,
    })
    # timestamps as proper Firestore timestamps
    plan_fields["submittedAt"] = {"timestampValue": now_iso}
    plan_fields["reviewedAt"] = {"timestampValue": now_iso}

    # Find an existing plan for this member+week to update in place.
    existing = run_query(f"families/{family_id}", "plans",
                         [("memberId", member_id), ("weekStart", week_start)],
                         token)
    if existing:
        doc_name = existing[0][0]
        rel = doc_name.split("/documents/")[-1]
        r = requests.patch(
            f"{FS_BASE}/{rel}",
            headers=auth_headers(token),
            json={"fields": plan_fields},
            timeout=30,
        )
        r.raise_for_status()
        print(f"Updated existing plan: {rel.split('/')[-1]}")
    else:
        r = requests.post(
            f"{FS_BASE}/families/{family_id}/plans",
            headers=auth_headers(token),
            json={"fields": plan_fields},
            timeout=30,
        )
        if r.status_code not in (200, 201):
            sys.exit(f"Create failed: {r.status_code} {r.text}")
        print(f"Created plan: {r.json()['name'].split('/')[-1]}")

    print("Done! Open the app → Planner for the week of "
          f"{week_start} to see the imported activities.")


if __name__ == "__main__":
    main()
